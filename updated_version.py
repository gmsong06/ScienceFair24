# FORMAT DIAGNOSES DATA

import openpyxl
from openpyxl.styles import Font


def extract_data(file_path):
    workbook = openpyxl.load_workbook(file_path)

    # Select first sheet
    sheet = workbook[workbook.sheetnames[0]]

    extracted_data = []
    years = {}
    max_years = 0


    for row in sheet.iter_rows(min_row=2, values_only=True):
        OASISID = row[1]
        years_to_visit = row[2]
        DementedLetter = row[3]
        CDR = row[4]

        if OASISID not in years:
            years[OASISID] = {
                "first_year": None,
                "last_year": None
            }

        if DementedLetter == 'MCI' and years[OASISID]["first_year"] is None or DementedLetter == 'DMN' and years[OASISID]["first_year"] is None:
            if years[OASISID]["first_year"] is None:
                years[OASISID]["first_year"] = years_to_visit - 1

        years[OASISID]["last_year"] = years_to_visit

        if years[OASISID]["first_year"] is not None:
            data = {
                "OASISID": OASISID,
                "years_to_visit": years_to_visit - years[OASISID]["first_year"],
                "DementedLetter": DementedLetter,
                "CDR": CDR
            }
            max_years = max(max_years, data["years_to_visit"])
            extracted_data.append(data)

    return extracted_data, max_years


def format_data(extracted_data, max_years, output_file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Add header row
    header = ["OASISID"]

    for year in range(max_years + 1):
        header.append(year)

    sheet.append(header)

    # Apply bold font to the header row
    bold_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold_font

    oasisid_data = {}  # Dictionary to store data for each OASISID

    for data in extracted_data:
        oasisid = data["OASISID"]
        year = data["years_to_visit"]
        cdr = data["CDR"]

        if oasisid in oasisid_data:
            oasisid_data[oasisid][year] = cdr
        else:
            oasisid_data[oasisid] = [None] * (max_years + 1)
            oasisid_data[oasisid][year] = cdr

    for oasisid, data_row in oasisid_data.items():
        row_data = [oasisid] + data_row
        sheet.append(row_data)

    workbook.save(output_file_path)
    print("Data formatted and saved to new Excel file:", output_file_path)


if __name__ == "__main__":
    data, max_years = extract_data("diagnosis_CDR.xlsx")
    # print(data)
    format_data(data, max_years, "formatted_diagnosis_CDR.xlsx")

